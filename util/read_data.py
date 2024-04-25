# -*- coding: utf-8 -*-
"""
@Time ： 2023/03/10 10:41
@Auth ： 章豹
@File ：read_data.py
@IDE ：PyCharm

"""
import csv
import os
from openpyxl import load_workbook
dir_path = os.path.dirname(__file__).split(sep="util")


def read_csv(filename):
    """
    读取csv文件内容并且返回结果为list
    :param filename: 文件所在的相对路径
    :return:
    """
    path = dir_path[0]+filename
    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for i in reader:  # 这边的i内容是字典形式的
            yield list(i.values())


def read_excel(filename):
    """
    读取excel的数据
    :param filename: 文件所在的相对路径
    :return:
    """
    path = dir_path[0]+filename
    wb = load_workbook(path)
    ws = wb.active
    yield from ws.iter_rows(min_row=2, values_only=True)
